#!/usr/bin/env python3
"""OPTIONAL: build an employer sponsorship index from DOL disclosure data.

This enriches scoring with real H-1B (LCA) and PERM (green-card) filing history
per employer. It is NOT required for the core search — run it once per quarter
if you want employer-level sponsorship signal.

Requires openpyxl (`.venv/bin/python -m pip install openpyxl`) and the DOL files.

Download the quarterly XLSX from the DOL OFLC performance-data page
(https://www.dol.gov/agencies/eta/foreign-labor/performance), e.g.:
  LCA_Disclosure_Data_FY2025_Q4.xlsx   (H-1B / H-1B1 / E-3)
  PERM_Disclosure_Data_FY2025.xlsx     (green card / PERM)

Usage:
  .venv/bin/python .agents/skills/job-search/scripts/build_sponsor_index.py \
      --lca path/to/LCA_Disclosure_Data_FY2025_Q4.xlsx \
      --perm path/to/PERM_Disclosure_Data_FY2025.xlsx \
      --out .agents/skills/job-search/data/sponsors.json

The resulting sponsors.json maps normalized-employer -> {"h1b": n, "perm": n}
and is auto-loaded by search_jobs.py when present.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from scoring import _norm_company

SKILL_DIR = Path(__file__).resolve().parents[1]


def _employer_col(header: list[str]) -> int | None:
    for i, h in enumerate(header):
        if h and "EMPLOYER" in str(h).upper() and "NAME" in str(h).upper():
            return i
    return None


def count_employers(path: str) -> dict[str, int]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl not installed. Run: .venv/bin/python -m pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows)]
    col = _employer_col(header)
    if col is None:
        sys.exit(f"Could not find an EMPLOYER_NAME column in {path}: {header[:8]}...")
    counts: dict[str, int] = {}
    for row in rows:
        if col < len(row) and row[col]:
            key = _norm_company(str(row[col]))
            if key:
                counts[key] = counts.get(key, 0) + 1
    wb.close()
    return counts


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--lca", help="Path to LCA (H-1B) disclosure XLSX")
    ap.add_argument("--perm", help="Path to PERM disclosure XLSX")
    ap.add_argument("--out", default=str(SKILL_DIR / "data" / "sponsors.json"))
    args = ap.parse_args()
    if not args.lca and not args.perm:
        sys.exit("Provide at least one of --lca / --perm.")

    index: dict[str, dict] = {}
    if args.lca:
        print("Parsing LCA (H-1B)...", file=sys.stderr)
        for emp, n in count_employers(args.lca).items():
            index.setdefault(emp, {"h1b": 0, "perm": 0})["h1b"] += n
    if args.perm:
        print("Parsing PERM (green card)...", file=sys.stderr)
        for emp, n in count_employers(args.perm).items():
            index.setdefault(emp, {"h1b": 0, "perm": 0})["perm"] += n

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(index, indent=0))
    print(f"Wrote {len(index)} employers -> {out}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
